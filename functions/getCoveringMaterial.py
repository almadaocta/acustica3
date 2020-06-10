from openpyxl import load_workbook
import openpyxl

def getCoveringMaterial(C):
    #Lectura de excel de materiales y obtencion de parametros del recubrimiento seleccionado
    materials = load_workbook(filename = 'resources/sheets/recubrimientos.xlsx')
    sheet_obj = materials.active 


    material=int(C)
    for cell in sheet_obj['A']:
        if material == cell.value:
            materialRow= cell.row
    R=[]
    for i in range(0,18):
        R.append(sheet_obj.cell(row = materialRow, column = i+7).value)

  
    
    return R