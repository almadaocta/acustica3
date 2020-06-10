from openpyxl import load_workbook
import openpyxl
import math
from math import *

def set_layer(calcData,thickness,material):


    materials = load_workbook(filename = 'resources/sheets/materials.xlsx')
    sheet_obj = materials.active 

    for cell in sheet_obj['C']:
        if(cell.value is not None):
            if material == cell.value:
                materialR= cell.row
                materialColumn= cell.column
        

    calcData['material2'] = material
    calcData['density2'] = (sheet_obj.cell(row = materialR, column = (cell.column + 1)).value)
    calcData['young2'] = (sheet_obj.cell(row = materialR, column = (cell.column + 2)).value)
    calcData['lossFactor2'] = (sheet_obj.cell(row = materialR, column = (cell.column + 3)).value)
    calcData['poisson2'] = (sheet_obj.cell(row = materialR, column = (cell.column + 4)).value)
    calcData['thickness2'] = thickness

    c = 343 #Velocidad del sonido en el aire
    p0 = 1.18 #Densidad del aire
    ly = calcData['height'] #Alto
    lx = calcData['length'] #Ancho
    E2=calcData['young2']
    E1=calcData['young']
    t1=calcData['thickness']
    t2=calcData['thickness2']
    O2=calcData['poisson2']
    O1=calcData['poisson']
    p2=calcData['density2']
    sd2 = p2*t2 #Densidad superficial material 2
    sdT = sd2 + calcData['sd'] #Densidad superficial total
    Y=((E1*(t1/2))+E2*(t1+(t2/2)))/(E1+E2)
    B1=calcData['B']
    B2=(E2/(1-O2**2))*((t2**3)/12)
    Beff=((E1*t1)/(12*(1-O1**2)))*(t1**2+12*((Y-(t1/2))**2))+((E2*t2)/(12*(1-O2**2)))*(t2**2+12*((Y-((2*t1+t2)/2))**2))

    if t1==t2 and calcData['material2']==calcData['material']:
        FcEQ=((c**2)/(2*pi))*sqrt(sdT/(B1+B2))
    else:    
        FcEQ=((c**2)/(2*pi))*sqrt(sdT/Beff)

    
    f11 = (c**2/(4*FcEQ))*((1/lx**2)+(1/ly**2)) #Modo (1,1) de placa del elemento 


    calcData['Fc'] = FcEQ
    calcData['sd'] = sdT
    calcData['f11'] = f11 
   
    