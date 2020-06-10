from openpyxl import load_workbook
import openpyxl
from os.path import dirname, abspath
from saveExport import export_save
from openpyxl.chart import Reference, Series, LineChart
import os.path
import string
import random

export={}
def randomString(stringLength=4):
    letters = string.ascii_lowercase
    return ''.join(random.choice(letters) for i in range(stringLength))
def set_export(rooms,structures,flanks):

    name=randomString()
    parent=dirname(dirname(abspath(__file__)))
    path = parent + '/export/' + name + '.xlsx' 
    

    export['file'] = load_workbook(filename = 'resources/sheets/exportTemplate.xlsx')
    sheets = export['file'].sheetnames
    wb = export['file'] 
    ws=wb[sheets[0]]


    #Masas superficiales 

    ws.cell(row = 6, column = 2).value = structures['D'].sd
    ws.cell(row = 6, column = 3).value = structures['L1'].sd
    ws.cell(row = 6, column = 4).value = structures['L2'].sd
    ws.cell(row = 6, column = 5).value = structures['T'].sd
    ws.cell(row = 6, column = 6).value = structures['P'].sd


    for i in range(0,20):
        ws.cell(row = 9+i, column = 2).value = structures['D'].R[i+4]
        ws.cell(row = 9+i, column = 3).value = structures['L1'].R[i+4]
        ws.cell(row = 9+i, column = 4).value = structures['L2'].R[i+4]
        ws.cell(row = 9+i, column = 5).value = structures['T'].R[i+4]
        ws.cell(row = 9+i, column = 6).value = structures['P'].R[i+4]


    export['file'].save(path)

    return name

