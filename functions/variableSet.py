
from math import *
import math
from openpyxl import load_workbook
import openpyxl

variables={}

def set_variables(structure,rooms):

    #Medidas estructura

        
    if structure.ID=='T':
        lx=rooms['E'].T
        ly=rooms['E'].L
    if structure.ID=='P':
        lx=rooms['E'].T
        ly=rooms['E'].L
    if structure.ID=='D':
        lx=rooms['E'].T
        ly=rooms['E'].H
    if structure.ID=='L1':
        lx=rooms['E'].L
        ly=rooms['E'].H
    if structure.ID=='L2':  
        lx=rooms['E'].L
        ly=rooms['E'].H             

    #Definicion de variables globales

    variables['c'] = 343 #Velocidad del sonido en el aire
    variables['p0'] = 1.18 #Densidad del aire
    variables['filterVector'] = [20,25,31.5,40,50,63,80,100,125,160,200,250,315,400,500,630,800,1000,1250,1600,2000,2500,3150,4000,5000,6300,8000,10000,12500,16000,20000]
    variables['db'] = 0.236
    variables['octave'] = 3   


    #Lectura de excel de materiales y obtencion de parametros del material seleccionado
    materials = load_workbook(filename = 'resources/sheets/materials.xlsx')
    sheet_obj = materials.active 


    material1=structure.M1
    for cell in sheet_obj['C']:
        if(cell.value is not None):
            if material1 == cell.value:
                materialR= cell.row
                materialColumn= cell.column
        
    structure.density = (sheet_obj.cell(row = materialR, column = (cell.column + 1)).value)
    structure.young = (sheet_obj.cell(row = materialR, column = (cell.column + 2)).value)
    structure.lossfactor = (sheet_obj.cell(row = materialR, column = (cell.column + 3)).value)
    structure.poisson = (sheet_obj.cell(row = materialR, column = (cell.column + 4)).value)
    if structure.T2>0:
        material2=structure.M1
        for cell in sheet_obj['C']:
            if(cell.value is not None):
                if material2 == cell.value:
                    materialR= cell.row
                    materialColumn= cell.column
            
        structure.density2 = (sheet_obj.cell(row = materialR, column = (cell.column + 1)).value)
        structure.young2 = (sheet_obj.cell(row = materialR, column = (cell.column + 2)).value)
        structure.lossfactor2 = (sheet_obj.cell(row = materialR, column = (cell.column + 3)).value)
        structure.poisson2 = (sheet_obj.cell(row = materialR, column = (cell.column + 4)).value)

    
    structure.sd1=structure.density*structure.T1
    structure.s= lx*ly
    structure.B=(structure.young/(1-structure.poisson**2))*((structure.T1**3)/12)
    structure.Fc=round((variables['c']**2/(2*math.pi))*(sqrt(structure.sd1/structure.B)),2)

    #Si son dos capas, calcular valores equivalentes 

    if structure.T2>0:
        E1=structure.young
        T1=structure.T1
        E2=structure.young2
        T2=structure.T2
        O1=structure.poisson
        O2=structure.poisson2
        B1=structure.B
        structure.sd2=structure.density2*structure.T2
        structure.sd=structure.sd1+structure.sd2
        Y=((E1*(T1/2))+E2*(T1+(T2/2)))/(E1+E2)
        B2=(E2/(1-O2**2))*((T2**3)/12)
        Beff=((E1*T1)/(12*(1-O1**2)))*(T1**2+12*((Y-(T1/2))**2))+((E2*T2)/(12*(1-O2**2)))*(T2**2+12*((Y-((2*T1+T2)/2))**2))
        if T1==T2 and structure.M1==structure.M2:
            structure.Fc=((variables['c']**2)/(2*pi))*sqrt(structure.sd/(B1+B2))
        else:    
            structure.Fc=((variables['c']**2)/(2*pi))*sqrt(structure.sd/Beff)
    else:
        structure.sd=structure.sd1
    
    structure.F11=(variables['c']**2/(4*structure.Fc))*((1/lx**2)+(1/ly**2))

  
    
    return variables

    

