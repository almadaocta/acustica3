from openpyxl import load_workbook
import openpyxl

def getStructureMaterial(structure):

    #Lectura de excel de materiales y obtencion de parametros del material seleccionado
    materials = load_workbook(filename = 'resources/sheets/estructura.xlsx')
    sheet_obj = materials.active 


    material=int(structure.M)
    for cell in sheet_obj['A']:
        if material == cell.value:
            materialRow= cell.row
    R=[]
    for i in range(0,18):
        R.append(sheet_obj.cell(row = materialRow, column = i+7).value)

    SD=sheet_obj.cell(row = materialRow, column = 6).value
    
    return R,SD